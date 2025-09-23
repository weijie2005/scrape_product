# Author: shen weijie
# Date: 2025-08-22
#----------------------

import time
import random
import os
import requests
import uuid
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from io import BytesIO

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup


# 配置常量
MAX_WAIT_TIME = 20
SCROLL_STEP = 5
SCROLL_DELAY = 0.1

# 新增：图片保存目录
IMG_SAVE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'product_images')
if not os.path.exists(IMG_SAVE_DIR):
    os.makedirs(IMG_SAVE_DIR)


# 下载图片函数
def download_image(image_url):
    try:
        # 处理相对URL
        if image_url.startswith('//'):
            image_url = 'https:' + image_url
        elif not image_url.startswith('http'):
            return None

        # 发送请求下载图片
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(image_url, headers=headers, timeout=10)
        response.raise_for_status()

        # 检查图片格式，如果是WebP则转换为JPEG
        image_data = BytesIO(response.content)
        try:
            image_pic = PILImage.open(image_data)
            if image_pic.format == 'WEBP':
                # 转换为JPEG格式
                image_pic = image_pic.convert('RGB')
                # 重新生成BytesIO对象
                converted_image_data = BytesIO()
                image_pic.save(converted_image_data, format='JPEG')
                # 重置指针位置
                converted_image_data.seek(0)
                # 使用转换后的数据
                image_data = converted_image_data
        except Exception as e:
            print(f'图片格式检查失败: {str(e)}')
            # 如果无法识别格式，继续使用原始数据

        # 生成唯一文件名
        filename = f'{uuid.uuid4()}.jpg'
        filepath = os.path.join(IMG_SAVE_DIR, filename)

        # 保存图片
        with open(filepath, 'wb') as f:
            f.write(image_data.getvalue())
        return filepath
    except Exception as e:
        print(f'图片下载失败: {str(e)}')
        return None

def init_browser():
    """
    配置并初始化浏览器，设置开发者模式以绕过网站检测。
    :return: 配置好的浏览器对象
    """
    # 配置 Chrome 选项
    chrome_options = Options()
    #chrome_options.add_argument('--headless')

    chrome_options.accept_insecure_certs = True

    user_agent = '''Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'''
    chrome_options.add_argument(f'user-agent={user_agent}')

    # 获取当前脚本的绝对路径
    script_path = os.path.abspath(__file__)    
    # 获取脚本所在的目录
    current_directory = os.path.dirname(script_path)

    # 配置 Chrome 驱动路径
    chrome_driver = f'{current_directory}\\selenium_driver\\chromedriver.exe'

    # 配置 Chrome 服务
    service = Service(chrome_driver)

    # 返回一个 chrome 浏览器对象
    chrome_brower = webdriver.Chrome(service=service, options=chrome_options)        

    # 使用 JavaScript 关闭浏览器自动化检测机制
    chrome_brower.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
                """
    })

    # 最大化浏览器窗口,不设置最大窗口，有些底部元素会被遮挡，无法完成点击跳转
    chrome_brower.maximize_window()    
    
    print(f'浏览器已初始化完成')

    return chrome_brower

def simulate_scroll(browser):
    """
    模拟鼠标慢慢滚动以加载商品图片。
    :param browser: 浏览器对象
    """
    # 获取页面高度
    page_height = browser.execute_script("return document.body.scrollHeight")
    current_position = 0
    while current_position < page_height:
        next_position = current_position + SCROLL_STEP
        browser.execute_script(f"window.scrollTo(0, {next_position});")
        time.sleep(SCROLL_DELAY)
        current_position = next_position

def get_products(browser, page_num, ws):
    """
    获取对应页码下的所有商品信息，并保存到 Excel 工作表中。
    :param browser: 浏览器对象
    :param page_num: 页码
    :param ws: Excel 工作表对象
    """
    print(f"正在提取第{page_num}页的商品信息...")
    time.sleep(random.randint(3, 5))

    simulate_scroll(browser)
    time.sleep(random.randint(3, 5))

    result_html = browser.page_source
    soup = BeautifulSoup(result_html, "html.parser")

    # 提取所有商品的共同父元素
    divs = soup.find_all('div', class_='tbpc-col search-content-col tbpc-col-lg-12 tbpc-col-xl-12 tbpc-col-xxl-10 tbpc-col tbpc-col-horizon-8 search-content-col tbpc-col-lg-12 tbpc-col-xl-12 tbpc-col-xxl-10')

    for div in divs:
        # 查找并获取需要的相关信息, 并进行空值检查
        product_url_tag = div.find('a', class_='doubleCardWrapperAdapt--mEcC7olq')
        if not product_url_tag:
            print(f"警告：未找到商品链接元素，跳过此商品")
            continue
        product_url = product_url_tag.attrs.get('href', '')
        
        image_tag = div.find('img', class_='mainImg--sPh_U37m')
        image_url = image_tag.attrs.get('src', '') if image_tag else ''
                        
        title_tag = div.find('div', class_='title--qJ7Xg_90')
        title = title_tag.find('span').text if (title_tag and title_tag.find('span')) else ''
        
        price_tag = div.find('div', class_='priceInt--yqqZMJ5a')
        price = price_tag.text if price_tag else ''
        
        deal_tag = div.find('span', class_='realSales--XZJiepmt')
        deal = deal_tag.text if deal_tag else '' 
        
        shop_name_tag = div.find('span', class_='shopNameText--DmtlsDKm')
        shop_name = shop_name_tag.text if shop_name_tag else ''
        
        shop_url_tag = div.find('a', class_='shopName--hdF527QA')
        shop_url = shop_url_tag.attrs.get('href', '') if shop_url_tag else ''
        
        location_tag = div.find('div', class_='procity--wlcT2xH9')
        location = location_tag.text if location_tag else ''

        product = {
            '商品图片': image_url,
            '商品网址': product_url,
            '价格': price,
            '商品简介': title,
            '交易数量': deal,
            '店铺名称': shop_name,
            '店铺网址': shop_url,
            '店铺所在地': location
        }

        print(f'{product}')

        if product:
            # 下载图片并插入Excel
            img_path = download_image(product['商品图片'])
            
            # 先添加数据行（图片列留空）
            current_row = ws.max_row + 1
            ws.append(['', product['商品网址'], product['价格'], product['商品简介'], product['交易数量'], product['店铺名称'], product['店铺网址'], product['店铺所在地']])
            
            # 插入图片到A列
            if img_path:
                img = Image(img_path)
                # 调整图片大小
                img.width = 100
                img.height = 100
                # 插入图片到当前行A列
                ws.add_image(img, f'A{current_row}')
                # 调整行高适应图片
                ws.row_dimensions[current_row].height = 80
        else:
            print("没有找到商品")


def page_turning(browser, page_num, ws):
    """
    跳转到指定页码并获取该页商品信息。
    :param browser: 浏览器对象
    :param page_num: 目标页码
    :param ws: Excel 工作表对象
    """
    print(f'正在跳转至第{page_num}页')
    try:
        # 滚动到页面底部
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # 等待指定元素加载完成
        wait=WebDriverWait(browser,MAX_WAIT_TIME)
        page_box=wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/span[3]/input')))

        print('定位到页面底部的页码输入框成功')

        # 定位到页码输入框并输入目标页码
        page_box.clear()
        page_box.send_keys(page_num)

        # 点击跳转按钮 
        wait=WebDriverWait(browser,MAX_WAIT_TIME)
        button=wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/button[3]')))
        button.click()

        # 新增：等待页面加载完成（等待商品列表元素出现）
        wait=WebDriverWait(browser,MAX_WAIT_TIME)        
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/span[3]/input')))


        print("跳转页面成功")

        get_products(browser, page_num, ws)

    except TimeoutException:
        print(f"跳转超时，重新跳转，当前页码：{page_num}")
        page_turning(browser, page_num, ws)

def fetch_goods(browser, start_page, total_pages, ws, url, excel_file_name, keyword):
    """
    搜索商品并抓取指定页码范围的商品信息。
    :param browser: 浏览器对象
    :param start_page: 起始页码
    :param total_pages: 总页数
    :param ws: Excel 工作表对象
    :param url: 搜索页面的 URL
    :param excel_file_name: 保存数据的 Excel 文件名
    :param keyword: 搜索关键词
    """
    print(f'正在爬取第{start_page}页')
    
    try:
        browser.get(url)
        time.sleep(10)  # 等待手动扫码登录

        # 找到搜索输入框并输入关键词
        wait=WebDriverWait(browser,MAX_WAIT_TIME)
        input_box = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#q")))
        
        wait=WebDriverWait(browser,MAX_WAIT_TIME)
        search_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#J_SearchForm > div > div.search-button > button')))

        # 在搜索输入框中输入关键词
        input_box.send_keys(keyword)

        # 点击搜索按钮
        search_button.click()

        time.sleep(20)  # 等待搜索结果加载，如有滑块手动操作

        if start_page != 1:
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.randint(1, 3))

            # 定位到页码输入框并输入目标页码
            wait=WebDriverWait(browser,MAX_WAIT_TIME)
            page_box = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/span[3]/input')))

            print('定位到页面底部的页码输入框成功')
            page_box.clear()
            page_box.send_keys(start_page)
            page_box.send_keys('\n')

            # 点击跳转按钮 
            wait=WebDriverWait(browser,MAX_WAIT_TIME)
            button=wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/button[3]')))
            button.click()

            # 新增：等待页面加载完成（等待商品列表元素出现）
            wait=WebDriverWait(browser,MAX_WAIT_TIME)        
            wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="search-content-leftWrap"]/div[3]/div[4]/div/div/span[3]/input')))

        # 获取起始页商品信息
        get_products(browser, start_page, ws)

        # 保存当前页数据
        ws.parent.save(excel_file_name)

        # 遍历后续页码并跳转获取商品信息
        for i in range(start_page + 1, start_page + total_pages):
            page_turning(browser, i, ws)

            # 每页完成后保存
            ws.parent.save(excel_file_name)            

        print(f'已完成第{start_page}页到第{start_page + total_pages}页的商品信息获取')

    except TimeoutException as e:
        print("搜索商品超时，重新搜索", e)
        #fetch_goods(browser, start_page, total_pages, ws, url, excel_file_name, keyword)


if __name__ == '__main__':
    #查的页面数量
    page_start = 1
    page_all = 10

    #查找的网站url
    base_url = 'https://s.taobao.com'

    # 搜索的关键词
    keyword = "水杯" 

    # 获取当前脚本的绝对路径
    script_path = os.path.abspath(__file__)    
    # 获取脚本所在的目录
    current_directory = os.path.dirname(script_path)    
    #保存文件名
    excel_file = f'{current_directory}/output/fetch_taobao_2025.xlsx'      
    
    # 初始化浏览器
    browser = init_browser()

    # 创建 Excel 工作簿和工作表
    wb = Workbook()
    ws = wb.active

    #设置工作表的列宽，并设置标题
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 100
    ws.column_dimensions['H'].width = 10

    # 设置自动换行
    wrap_alignment = Alignment(wrap_text=True)

    # 为需要自动换行的列设置格式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap_alignment

    ws.append(['商品图片', '商品网址', '价格', '商品简介', '交易数量', '店铺名称', '店铺网址', '店铺所在地'])

    # 开始搜索商品并抓取数据
    fetch_goods(browser, page_start, page_all, ws, base_url, excel_file,keyword)     

    # 关闭浏览器
    #browser.quit()